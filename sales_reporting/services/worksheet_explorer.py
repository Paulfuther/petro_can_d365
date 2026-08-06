from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal
from pathlib import Path
from typing import Any, List

from openpyxl import load_workbook


class WorksheetExplorerError(Exception):
    """Raised when a workbook cannot be inspected."""


@dataclass(frozen=True)
class WorksheetCell:
    coordinate: str
    column: str
    value: Any


@dataclass(frozen=True)
class WorksheetRow:
    row_number: int
    cells: List[WorksheetCell]


@dataclass(frozen=True)
class WorksheetInspection:
    source_filename: str
    worksheet_name: str
    rows: List[WorksheetRow]


def _display_value(value):
    """
    Return a template-friendly representation of an Excel value.
    """
    if value is None:
        return ""

    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M:%S")

    if isinstance(value, date):
        return value.isoformat()

    if isinstance(value, Decimal):
        return format(value, "f")

    return value


def inspect_worksheet(
    uploaded_file,
    *,
    max_rows: int = 150,
    min_column: int = 1,
    max_column: int = 9,
) -> WorksheetInspection:
    """
    Inspect columns A through I of the first worksheet.

    Empty rows are omitted. No database changes are made.
    """
    filename = Path(uploaded_file.name).name

    try:
        uploaded_file.seek(0)

        workbook = load_workbook(
            uploaded_file,
            data_only=True,
            read_only=True,
        )
    except Exception as exc:
        raise WorksheetExplorerError(
            f"Could not open {filename}: {exc}"
        ) from exc

    try:
        if not workbook.sheetnames:
            raise WorksheetExplorerError(
                f"{filename} contains no worksheets."
            )

        worksheet = workbook[workbook.sheetnames[0]]
        inspected_rows = []

        for row_number, row in enumerate(
            worksheet.iter_rows(
                min_row=1,
                max_row=max_rows,
                min_col=min_column,
                max_col=max_column,
            ),
            start=1,
        ):
            populated_cells = []

            for cell in row:
                if cell.value in (None, ""):
                    continue

                populated_cells.append(
                    WorksheetCell(
                        coordinate=cell.coordinate,
                        column=cell.column_letter,
                        value=_display_value(cell.value),
                    )
                )

            if populated_cells:
                inspected_rows.append(
                    WorksheetRow(
                        row_number=row_number,
                        cells=populated_cells,
                    )
                )

        return WorksheetInspection(
            source_filename=filename,
            worksheet_name=worksheet.title,
            rows=inspected_rows,
        )

    finally:
        workbook.close()
        uploaded_file.seek(0)