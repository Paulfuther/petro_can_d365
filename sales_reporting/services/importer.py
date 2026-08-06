import re
from dataclasses import dataclass
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Optional
from django.db import transaction
from openpyxl import load_workbook

from sales_reporting.models import (
    Category,
    CategoryResult,
    SalesImport,
    Store,
)

from .workbook_reader import (
    WorkbookMetadata,
    WorkbookReadError,
    read_workbook_metadata,
)

CATEGORY_PATTERN = re.compile(
    r"^\s*(?P<code>\d+)\s*-\s*(?P<name>.+?)\s*$"
)


class SalesImportError(Exception):
    """Raised when a D365 sales workbook cannot be imported."""


@dataclass(frozen=True)
class ParsedCategoryResult:
    code: str
    name: str
    level: int
    parent_code: Optional[str]
    quantity: Decimal
    sales: Decimal
    cogs: Decimal
    gross_margin: Decimal
    margin_percentage: Decimal


@dataclass(frozen=True)
class SalesImportSummary:
    source_filename: str
    store_number: str
    reporting_month: object
    categories_created: int
    categories_updated: int
    category_results_created: int
    replaced_existing_import: bool

def _to_decimal(value, field_name: str, row_number: int) -> Decimal:
    if value in (None, ""):
        return Decimal(0)

    if isinstance(value, Decimal):
        return value

    try:
        return Decimal(str(value))
    except (InvalidOperation, TypeError, ValueError) as exc:
        raise SalesImportError(
            f"Row {row_number}: could not read {field_name} "
            f"value {value!r}."
        ) from exc

def _category_level_from_cell(cell, code: str) -> int:
    """
    Determine the D365 hierarchy level from the category code.

    Examples:

        130   Convenience Store       level 1
        2320  Alternate Beverages     level 2
        31280 Energy Drinks           level 3
        42260 <=600ML                 level 4

    D365 temporary categories use:

        99910 Beverages               level 3
        98010 <600 ML                 level 4
    """
    if len(code) == 3:
        return 1

    if len(code) == 4:
        return 2

    if len(code) == 5:
        if code.startswith("3"):
            return 3

        if code.startswith("999"):
            return 3

        return 4

    raise SalesImportError(
        f"Unsupported D365 category code: {code}"
    )

def _parse_category_row(row, row_number: int):
    """
    Return a ParsedCategoryResult when the row is a category result.

    Expected columns:
        A = code and category name
        C = quantity
        D = sales
        E = COGS
        F = gross margin
        I = margin percentage
    """
    category_cell = row[0]
    category_text = category_cell.value

    if category_text in (None, ""):
        return None

    category_text = str(category_text).strip()
    match = CATEGORY_PATTERN.match(category_text)

    if not match:
        return None

    measured_values = (
        row[2].value,
        row[3].value,
        row[4].value,
        row[5].value,
        row[8].value,
    )

    if all(
        value in (None, "")
        for value in measured_values
    ):
        return None

    code = match.group("code")
    name = match.group("name").strip()

    return ParsedCategoryResult(
        code=code,
        name=name,
        level=_category_level_from_cell(
            category_cell,
            code,
        ),
        parent_code=None,
        quantity=_to_decimal(
            row[2].value,
            "quantity",
            row_number,
        ),
        sales=_to_decimal(
            row[3].value,
            "sales",
            row_number,
        ),
        cogs=_to_decimal(
            row[4].value,
            "COGS",
            row_number,
        ),
        gross_margin=_to_decimal(
            row[5].value,
            "gross margin",
            row_number,
        ),
        margin_percentage=_to_decimal(
            row[8].value,
            "margin percentage",
            row_number,
        ),
    )

def parse_category_results(
    uploaded_file,
) -> list[ParsedCategoryResult]:
    filename = Path(uploaded_file.name).name

    try:
        uploaded_file.seek(0)

        workbook = load_workbook(
            uploaded_file,
            data_only=True,
            read_only=True,
        )
    except Exception as exc:
        raise SalesImportError(
            f"Could not open {filename}: {exc}"
        ) from exc

    try:
        if not workbook.sheetnames:
            raise SalesImportError(
                f"{filename} contains no worksheets."
            )

        worksheet = workbook[
            workbook.sheetnames[0]
        ]

        parsed_results = []
        latest_code_by_level = {}

        for row_number, row in enumerate(
            worksheet.iter_rows(
                min_row=1,
                min_col=1,
                max_col=9,
            ),
            start=1,
        ):
            parsed = _parse_category_row(
                row,
                row_number,
            )

            if parsed is None:
                continue

            parent_code = latest_code_by_level.get(
                parsed.level - 1
            )

            parsed = ParsedCategoryResult(
                code=parsed.code,
                name=parsed.name,
                level=parsed.level,
                parent_code=parent_code,
                quantity=parsed.quantity,
                sales=parsed.sales,
                cogs=parsed.cogs,
                gross_margin=parsed.gross_margin,
                margin_percentage=(
                    parsed.margin_percentage
                ),
            )

            parsed_results.append(parsed)

            latest_code_by_level[
                parsed.level
            ] = parsed.code

            deeper_levels = [
                level
                for level in latest_code_by_level
                if level > parsed.level
            ]

            for level in deeper_levels:
                del latest_code_by_level[level]

        if not parsed_results:
            raise SalesImportError(
                f"No category result rows were found "
                f"in {filename}."
            )

        return parsed_results

    finally:
        workbook.close()
        uploaded_file.seek(0)

def import_sales_workbook(
    uploaded_file,
    *,
    replace_existing: bool = False,
) -> SalesImportSummary:
    """
    Import one D365 category sales workbook.

    The entire replacement is atomic. If any row fails, the original
    store-month import remains untouched.
    """
    try:
        metadata: WorkbookMetadata = read_workbook_metadata(
            uploaded_file
        )
    except WorkbookReadError as exc:
        raise SalesImportError(str(exc)) from exc

    category_rows = parse_category_results(uploaded_file)

    with transaction.atomic():
        store, _ = Store.objects.update_or_create(
            number=metadata.store_number,
            defaults={
                "name": metadata.store_name,
            },
        )

        existing_import = SalesImport.objects.filter(
            store=store,
            reporting_month=metadata.reporting_month,
        ).first()

        replaced_existing_import = False

        if existing_import:
            if not replace_existing:
                raise SalesImportError(
                    f"Store {store.number} already has an import "
                    f"for {metadata.reporting_month:%B %Y}."
                )

            existing_import.delete()
            replaced_existing_import = True

        sales_import = SalesImport.objects.create(
            source_filename=metadata.source_filename,
            store=store,
            from_date=metadata.from_date,
            to_date=metadata.to_date,
            reporting_month=metadata.reporting_month,
            channel=metadata.channel,
        )

        categories_created = 0
        categories_updated = 0
        results_to_create = []

        categories_by_code = {}

        for parsed in category_rows:
            parent = None

            if parsed.parent_code:
                parent = categories_by_code.get(
                    parsed.parent_code
                )

                if parent is None:
                    parent = Category.objects.filter(
                        code=parsed.parent_code
                    ).first()

            category, created = Category.objects.get_or_create(
                code=parsed.code,
                defaults={
                    "name": parsed.name,
                    "level": parsed.level,
                    "parent": parent,
                },
            )

            if created:
                categories_created += 1
            else:
                changed_fields = []

                if category.name != parsed.name:
                    category.name = parsed.name
                    changed_fields.append("name")

                if category.level != parsed.level:
                    category.level = parsed.level
                    changed_fields.append("level")

                expected_parent_id = (
                    parent.id
                    if parent
                    else None
                )

                if category.parent_id != expected_parent_id:
                    category.parent = parent
                    changed_fields.append("parent")

                if changed_fields:
                    category.save(
                        update_fields=changed_fields
                    )
                    categories_updated += 1

            categories_by_code[parsed.code] = category

            results_to_create.append(
                CategoryResult(
                    sales_import=sales_import,
                    category=category,
                    quantity=parsed.quantity,
                    sales=parsed.sales,
                    cogs=parsed.cogs,
                    gross_margin=parsed.gross_margin,
                    margin_percentage=(
                        parsed.margin_percentage
                    ),
                )
            )

        CategoryResult.objects.bulk_create(
            results_to_create,
            batch_size=500,
        )

        if hasattr(sales_import, "status"):
            sales_import.status = "Completed"
            sales_import.save(update_fields=["status"])

    return SalesImportSummary(
        source_filename=metadata.source_filename,
        store_number=metadata.store_number,
        reporting_month=metadata.reporting_month,
        categories_created=categories_created,
        categories_updated=categories_updated,
        category_results_created=len(results_to_create),
        replaced_existing_import=replaced_existing_import,
    )