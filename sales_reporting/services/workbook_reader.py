import re
from calendar import monthrange
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Optional

from openpyxl import load_workbook


class WorkbookReadError(Exception):
    """Raised when a D365 workbook cannot be read or validated."""


@dataclass(frozen=True)
class WorkbookMetadata:
    source_filename: str
    worksheet_name: str
    store_number: str
    store_name: str
    from_date: date
    to_date: date
    reporting_month: date
    channel: str


STORE_PATTERN = re.compile(
    r"^\s*(?P<store>\d+)\s*-\s*(?P<name>.+?)\s*$"
)


def _normalize_text(value) -> str:
    if value is None:
        return ""

    return str(value).strip()


def _parse_date(value, label: str) -> date:
    if isinstance(value, datetime):
        return value.date()

    if isinstance(value, date):
        return value

    text = _normalize_text(value)

    if not text:
        raise WorkbookReadError(f"{label} is blank.")

    accepted_formats = (
        "%Y-%m-%d",
        "%Y/%m/%d",
        "%m/%d/%Y",
        "%d/%m/%Y",
    )

    for date_format in accepted_formats:
        try:
            return datetime.strptime(text, date_format).date()
        except ValueError:
            continue

    raise WorkbookReadError(
        f"Could not understand {label}: {text!r}"
    )


def _find_label_value(
    worksheet,
    label: str,
    max_rows: int = 25,
    max_columns: int = 12,
):
    """
    Find a label such as 'From date' and return the next non-empty
    value to its right.
    """
    target = label.casefold()

    for row in worksheet.iter_rows(
        min_row=1,
        max_row=max_rows,
        min_col=1,
        max_col=max_columns,
    ):
        for index, cell in enumerate(row):
            cell_text = _normalize_text(cell.value).casefold()

            if cell_text != target:
                continue

            for following_cell in row[index + 1:]:
                if following_cell.value not in (None, ""):
                    return following_cell.value

            raise WorkbookReadError(
                f"Found {label!r}, but no value appeared beside it."
            )

    raise WorkbookReadError(
        f"Could not find {label!r} in the workbook header."
    )


def _find_store(
    worksheet,
    max_rows: int = 20,
    max_columns: int = 6,
) -> tuple[str, str]:
    """
    Find a header value such as:
        00396 - 1553690 ONTARIO INC.
    """
    for row in worksheet.iter_rows(
        min_row=1,
        max_row=max_rows,
        min_col=1,
        max_col=max_columns,
    ):
        for cell in row:
            text = _normalize_text(cell.value)

            if not text:
                continue

            match = STORE_PATTERN.match(text)

            if not match:
                continue

            store_number = match.group("store").zfill(5)
            store_name = match.group("name").strip()

            return store_number, store_name

    raise WorkbookReadError(
        "Could not find the store number and store name "
        "in the workbook header."
    )


def _validate_reporting_period(
    from_date: date,
    to_date: date,
) -> date:
    if from_date > to_date:
        raise WorkbookReadError(
            "The report's From date is later than its To date."
        )

    if (
        from_date.year != to_date.year
        or from_date.month != to_date.month
    ):
        raise WorkbookReadError(
            "The report crosses more than one calendar month."
        )

    last_day = monthrange(
        from_date.year,
        from_date.month,
    )[1]

    if from_date.day != 1 or to_date.day != last_day:
        raise WorkbookReadError(
            "The report does not cover a complete calendar month. "
            f"Found {from_date} through {to_date}."
        )

    return from_date.replace(day=1)


def read_workbook_metadata(uploaded_file) -> WorkbookMetadata:
    """
    Read and validate the header of one uploaded D365 XLSX report.

    This function does not write anything to the database.
    """
    filename = Path(uploaded_file.name).name

    try:
        uploaded_file.seek(0)

        workbook = load_workbook(
            uploaded_file,
            data_only=True,
            read_only=True,
        )
    except Exception as exc:
        raise WorkbookReadError(
            f"Could not open {filename}: {exc}"
        ) from exc

    try:
        if not workbook.sheetnames:
            raise WorkbookReadError(
                f"{filename} does not contain any worksheets."
            )

        worksheet = workbook[workbook.sheetnames[0]]

        store_number, store_name = _find_store(worksheet)

        from_date = _parse_date(
            _find_label_value(worksheet, "From date"),
            "From date",
        )

        to_date = _parse_date(
            _find_label_value(worksheet, "To date"),
            "To date",
        )

        channel = _normalize_text(
            _find_label_value(worksheet, "Channel")
        )

        reporting_month = _validate_reporting_period(
            from_date,
            to_date,
        )

        return WorkbookMetadata(
            source_filename=filename,
            worksheet_name=worksheet.title,
            store_number=store_number,
            store_name=store_name,
            from_date=from_date,
            to_date=to_date,
            reporting_month=reporting_month,
            channel=channel,
        )

    finally:
        workbook.close()
        uploaded_file.seek(0)